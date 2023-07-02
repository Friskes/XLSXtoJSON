# pip install openpyxl
from openpyxl import load_workbook
from json import dump, load, dumps
from SpinCursor import SpinCursor
from typing import List


spin = SpinCursor(
    msg="Ожидайте выполнения программы",
    del_msg_after_stop=True,
    maxspin=500,
    minspin=0,
    speed=2,
    animType='sticks'
)

AVAILABLE_SUBCLASSES = {
    # "4-Shirt": [
    #     "Miscellaneous" # Miscellaneous это разнообразные итемы, могут быть как оружием так и бронёй.
    # ],
    # "21-Main Hand": [
    #     "Sword",
    #     "Axe",
    #     "Mace",
    #     "Miscellaneous",
    #     "Dagger",
    #     "Fist Weapon",
    #     "Exotic"
    # ],
    # "7-Legs": [
    #     "Cloth",
    #     "Leather",
    #     "Mail",
    #     "Plate"
    # ],
    # "8-Feet": [
    #     "Miscellaneous",
    #     "Cloth",
    #     "Leather",
    #     "Mail",
    #     "Plate"
    # ],
    # "5-Chest": [
    #     "Leather",
    #     "Cloth",
    #     "Mail",
    #     "Miscellaneous",
    #     "Plate"
    # ],
    # "21-Two-Hand": [
    #     "Sword",
    #     "Axe",
    #     "Staff",
    #     "Mace",
    #     "Polearm",
    #     "Miscellaneous",
    #     "Fishing Pole"
    # ],
    # "10-Hands": [
    #     "Leather",
    #     "Cloth",
    #     "Mail",
    #     "Plate",
    #     "Miscellaneous"
    # ],
    # "21-One-Hand": [
    #     "Sword",
    #     "Mace",
    #     "Dagger",
    #     "Axe",
    #     "Fist Weapon",
    #     "Miscellaneous"
    # ],
    # "9-Wrist": [
    #     "Leather",
    #     "Mail",
    #     "Cloth",
    #     "Plate",
    #     "Miscellaneous"
    # ],
    # "6-Waist": [
    #     "Cloth",
    #     "Mail",
    #     "Leather",
    #     "Plate",
    #     "Miscellaneous"
    # ],
    # "22-Off Hand": [
    #     "Shield",
    #     "Axe",
    #     "Fist Weapon",
    #     "Sword",
    #     "Dagger",
    #     "Mace"
    # ],
    # "22-Held In Off-hand": [
    #     "Miscellaneous"
    # ],
    # "16-Back": [
    #     "Cloth"
    # ],
    # "1-Head": [
    #     "Mail",
    #     "Cloth",
    #     "Leather",
    #     "Plate",
    #     "Miscellaneous"
    # ],
    # "3-Shoulder": [
    #     "Mail",
    #     "Leather",
    #     "Cloth",
    #     "Plate",
    #     "Miscellaneous"
    # ],
    # "23-Ranged": [
    #     "Gun",
    #     "Bow",
    #     "Wand",
    #     "Crossbow"
    # ],
    # "19-Tabard": [
    #     "Miscellaneous"
    # ],
    # "23-Thrown": [
    #     "Thrown"
    # ]
}

ITEM_QUALITIES = {
    'Poor': 0,      # (gray)
    'Common': 1,    # (white)
    'Uncommon': 2,  # (green)
    'Rare': 3,      # (blue)
    'Epic': 4,      # (purple)
    'Legendary': 5, # (orange)
    'Heirloom': 7   # (blizzard blue) фамильные
}

ALLOWED_SLOTS = {
    'Shirt': 4,
    'Main Hand': 21,
    'Legs': 7,
    'Feet': 8,
    'Chest': 5,
    'Two-Hand': 21,
    'Hands': 10,
    'One-Hand': 21,
    # 'Trinket': 13, # 14
    'Wrist': 9,
    'Waist': 6,
    # 'Finger': 11, # 12
    'Off Hand': 22,
    'Held In Off-hand': 22,
    'Back': 16,
    'Head': 1,
    # 'Neck': 2,
    'Shoulder': 3,
    'Ranged': 23,
    # 'Non-equippable': None,
    'Tabard': 19,
    # 'Relic': 0,
    'Thrown': 23
}

data = [{
    1: [],
    3: [],
    4: [],
    5: [],
    6: [],
    7: [],
    8: [],
    9: [],
    10: [],
    16: [],
    19: [],
    21: [],
    22: [],
    23: [],
    'enchants': [],
    'mounts': []
}]

rows_recorded = 0


def append_items_to_data(medium_data: dict):

    for item_data in clean_data:

        displayId = medium_data.get(item_data['itemId'])
        if not displayId: continue

        ITEM_DATA = {
            'itemId': item_data['itemId'],
            'displayId': displayId,
            'type': item_data['type'],
            'quality': item_data['quality'],
            'ilvl': item_data['ilvl'],
            'icon': item_data['icon'],
            'name': item_data['name']
        }
        data[0][ALLOWED_SLOTS[item_data['slot']]].append(ITEM_DATA)

        global rows_recorded
        rows_recorded += 1


def read_data_from_xlsx():
    medium_data = {}
    for row in sheet.iter_rows(min_row=2):

        row_data = {}
        for cell in row:

            if cell.value is None: continue
            row_data.update({column_names[cell.column-1]: cell.value})

        if not row_data.get('displayInfoId'): continue
        medium_data.update({row_data['ItemID']: row_data['displayInfoId']})

    append_items_to_data(medium_data)


def sorting_data_by_quality_and_ilvl_or_name() -> List[dict]:
    return [{slot_id: sorted(item_data, reverse=True, key=lambda item: (int(item['quality']), int(item['ilvl'])))
             if isinstance(slot_id, int) else item_data # sorted(item_data, key=lambda item: item['name'])
             for slot_data in data for slot_id, item_data in slot_data.items()}]


def get_column_names() -> List[str]:
    # строка, колонка
    return [sheet[1][i].value for i in range(max_column)]


def write_data_to_json(file_name: str, data):
    with open(file_name, 'w', encoding='utf-8') as file:
        dump(data, file, ensure_ascii=False)


def read_data_from_json(file_name: str):
    with open(file_name, 'r', encoding='utf-8') as file:
        return load(file)


def get_clean_data(file_name: str) -> List[dict]:

    items_data: List[dict] = read_data_from_json(file_name)

    clean_data = []
    for item_data in items_data:

        if item_data['class'] != 'Armor' and item_data['class'] != 'Weapon': continue
        if item_data['slot'] not in ALLOWED_SLOTS: continue
        if not item_data.get('subclass'): continue

        slotId_and_slotName = f"{ALLOWED_SLOTS[item_data['slot']]}-{item_data['slot']}"

        if slotId_and_slotName not in AVAILABLE_SUBCLASSES:
            AVAILABLE_SUBCLASSES.update({slotId_and_slotName: []})
        else:
            if item_data['subclass'] not in AVAILABLE_SUBCLASSES[slotId_and_slotName]:
                AVAILABLE_SUBCLASSES[slotId_and_slotName].append(item_data['subclass'])

        clean_data.append({
            'itemId': item_data['itemId'],
            'type': item_data['subclass'],
            'quality': ITEM_QUALITIES[item_data['quality']],
            'ilvl': item_data['itemLevel'],
            'icon': item_data['icon'],
            'name': item_data['name'],
            'slot': item_data['slot']
        })

    return clean_data


def append_enchants_to_data(file_name: str):
    enchants_data: List[dict] = read_data_from_json(file_name)

    for enchant in enchants_data[0].values():
        ENCHANT_DATA = {
            'visualId': enchant['visual'],
            'name': enchant['name'],
            'icon': enchant['icon']
        }
        data[0]['enchants'].append(ENCHANT_DATA)


def append_mounts_to_data(file_name: str):
    mounts_data: List[dict] = read_data_from_json(file_name)

    # с (mountId +- 382) заканчивается WRATH и начинаются следующие патчи (элементы массива отсортированы по mountId)
    for mount in mounts_data[0].values():
        MOUNT_DATA = {
            'spellId': mount['id'],
            'displayId': mount['npcmodel'],
            'name': mount['name']
        }
        data[0]['mounts'].append(MOUNT_DATA)


if __name__ == '__main__':
    spin.start()

    wb = load_workbook('files/items.xlsx', read_only=True)
    sheet = wb[wb.sheetnames[0]]
    max_row = sheet.max_row
    max_column = sheet.max_column

    column_names = get_column_names()

    clean_data = get_clean_data('files/data.json')

    read_data_from_xlsx()

    append_enchants_to_data('files/enchants.json')
    append_mounts_to_data('files/mounts.json')

    sorted_data = sorting_data_by_quality_and_ilvl_or_name()

    write_data_to_json('files/itemsdata.json', sorted_data)

    spin.stop()
    spin.join()

    print('\nAVAILABLE_SUBCLASSES =', dumps(AVAILABLE_SUBCLASSES, indent=4, ensure_ascii=False))

    percent_diff = (rows_recorded * 100) / max_row
    print(f'\nИтемов было добавлено: {rows_recorded}\nОт их общего количества: {max_row}')
    print(f'\nПроцент добавленных итемов от их общего количества: {percent_diff:.0f}%')

    print('\nПрограмма успешно завершилась!\n')
